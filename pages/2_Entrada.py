import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 

wb = load_workbook('content\spent.xlsx')
ws = wb.active

st.set_page_config(layout="wide")

col1, col2, col3, col4 = st.columns(4)

with col1:
    gasto1 = st.text_input('Qual seu nome?', '')
with col2:
    gasto2 = st.text_input('Adicione o valor de entrada de dinheiro:', '')
with col3:
    gasto3 = st.text_input('Adicione a data da entrada:', '')
with col4:
    gasto4 = st.text_input('De onde veio?', '')

for row in range(1, 4):
    for col in range(1, 5):
        char = get_column_letter(col)
        st.write(ws[char + str(row)].value)